import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule

def create_gantt_chart():
    # Make sure output directory exists
    os.makedirs('projects/project-management-excel', exist_ok=True)
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # ---------------------------------------------------------
    # STYLING DEFINITIONS (Mocha Warm Theme)
    # ---------------------------------------------------------
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="6B4F3A")
    subtitle_font = Font(name=font_family, size=11, italic=True, color="8A6D55")
    section_font = Font(name=font_family, size=12, bold=True, color="6B4F3A")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    data_bold_font = Font(name=font_family, size=10, bold=True, color="333333")
    note_font = Font(name=font_family, size=9, italic=True, color="666666")
    
    # Fills
    primary_header_fill = PatternFill(start_color="6B4F3A", end_color="6B4F3A", fill_type="solid") # Dark Espresso
    sub_header_fill = PatternFill(start_color="A47864", end_color="A47864", fill_type="solid") # Soft Mocha
    accent_fill = PatternFill(start_color="E8C9A0", end_color="E8C9A0", fill_type="solid") # Sand Gold
    zebra_fill = PatternFill(start_color="F9F6F0", end_color="F9F6F0", fill_type="solid") # Light Warm Ivory
    card_fill = PatternFill(start_color="F5EDE3", end_color="F5EDE3", fill_type="solid") # Ivory
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D6C5B3")
    thick_border_side = Side(border_style="medium", color="6B4F3A")
    double_border_side = Side(border_style="double", color="6B4F3A")
    
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    bottom_heavy_border = Border(bottom=thick_border_side)
    double_bottom_border = Border(bottom=double_border_side, top=thin_border_side)
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thick_border_side, bottom=thick_border_side)
    
    # Alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_vertical_header = Alignment(horizontal="center", vertical="center", text_rotation=90)
    
    # ---------------------------------------------------------
    # SHEET 1: README & GUIDE
    # ---------------------------------------------------------
    ws_readme = wb.create_sheet(title="README & Guide")
    ws_readme.views.sheetView[0].showGridLines = True
    
    ws_readme.column_dimensions['A'].width = 3
    ws_readme.column_dimensions['B'].width = 25
    ws_readme.column_dimensions['C'].width = 50
    
    # Title
    ws_readme['B2'] = "GANTT CHART INTERAKTIF KELOMPOK"
    ws_readme['B2'].font = title_font
    ws_readme['B3'] = "Panduan Penggunaan Template untuk Koordinasi Tugas Kuliah"
    ws_readme['B3'].font = subtitle_font
    
    # Section 1: Cara Penggunaan
    ws_readme['B5'] = "Cara Penggunaan Template"
    ws_readme['B5'].font = section_font
    
    steps = [
        ("1. Set Tanggal Mulai", "Buka tab 'Gantt Chart Schedule' lalu isi tanggal mulai proyek di sel C4. Kalender sumbu waktu sebelah kanan akan bergeser otomatis."),
        ("2. Input Daftar Tugas", "Isi kolom Tugas (B12 dst) beserta PIC yang bertanggung jawab (C12 dst) pada tabel jadwal."),
        ("3. Input Tanggal Tugas", "Isi 'Start Date' dan 'End Date' tugas. Durasi (hari) akan dihitung otomatis oleh formula Excel."),
        ("4. Update Progress", "Masukkan persentase penyelesaian tugas di kolom '% Progress' (G12 dst, contoh: isi 0.5 untuk 50%). Status akan terupdate otomatis (Not Started, In Progress, Completed)."),
        ("5. Pantau Gantt Chart", "Gantt Chart di kolom J ke kanan akan terarsir secara otomatis berdasarkan rentang tanggal tugas."),
        ("6. Pantau Dashboard KPI", "Buka tab 'KPI Dashboard' untuk melihat perkembangan keseluruhan proyek, sisa tugas, dan beban kerja per anggota kelompok.")
    ]
    
    row_idx = 7
    for step_title, step_desc in steps:
        ws_readme.cell(row=row_idx, column=2, value=step_title).font = Font(name=font_family, size=10, bold=True, color="6B4F3A")
        ws_readme.cell(row=row_idx, column=2).alignment = align_left
        ws_readme.cell(row=row_idx, column=3, value=step_desc).font = data_font
        ws_readme.cell(row=row_idx, column=3).alignment = align_left
        ws_readme.cell(row=row_idx, column=2).border = Border(bottom=Side(style="dashed", color="E8C9A0"))
        ws_readme.cell(row=row_idx, column=3).border = Border(bottom=Side(style="dashed", color="E8C9A0"))
        row_idx += 1
        
    # Section 2: Shortcuts Esensial
    ws_readme.cell(row=row_idx+1, column=2, value="Shortcut Excel Produktif").font = section_font
    row_idx += 3
    
    shortcuts = [
        ("Ctrl + ;", "Memasukkan tanggal hari ini secara instan."),
        ("F4", "Mengunci rumus menjadi absolut (tanda $) saat mengedit formula."),
        ("Alt + Enter", "Membuat baris baru di dalam satu sel yang sama (wrap text manual)."),
        ("Ctrl + Z", "Membatalkan kesalahan edit terakhir (Undo).")
    ]
    
    # Header untuk tabel shortcuts
    ws_readme.cell(row=row_idx, column=2, value="Shortcut").font = header_font
    ws_readme.cell(row=row_idx, column=2).fill = primary_header_fill
    ws_readme.cell(row=row_idx, column=2).alignment = align_center
    ws_readme.cell(row=row_idx, column=3, value="Fungsi & Kegunaan").font = header_font
    ws_readme.cell(row=row_idx, column=3).fill = primary_header_fill
    ws_readme.cell(row=row_idx, column=3).alignment = align_center
    row_idx += 1
    
    for sc_key, sc_desc in shortcuts:
        c1 = ws_readme.cell(row=row_idx, column=2, value=sc_key)
        c1.font = data_bold_font
        c1.alignment = align_center
        c1.border = thin_border
        
        c2 = ws_readme.cell(row=row_idx, column=3, value=sc_desc)
        c2.font = data_font
        c2.alignment = align_left
        c2.border = thin_border
        
        row_idx += 1

    # ---------------------------------------------------------
    # SHEET 2: KPI DASHBOARD
    # ---------------------------------------------------------
    ws_dash = wb.create_sheet(title="KPI Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash.column_dimensions['A'].width = 3
    ws_dash.column_dimensions['B'].width = 18
    ws_dash.column_dimensions['C'].width = 18
    ws_dash.column_dimensions['D'].width = 5
    ws_dash.column_dimensions['E'].width = 18
    ws_dash.column_dimensions['F'].width = 18
    ws_dash.column_dimensions['G'].width = 5
    ws_dash.column_dimensions['H'].width = 25
    ws_dash.column_dimensions['I'].width = 15
    ws_dash.column_dimensions['J'].width = 18
    
    # Title
    ws_dash['B2'] = "DASHBOARD PERKEMBANGAN PROYEK"
    ws_dash['B2'].font = title_font
    ws_dash['B3'] = "Rekapitulasi Kinerja Tugas Kelompok secara Real-Time"
    ws_dash['B3'].font = subtitle_font
    
    # KPI Cards (Row 5 - 7)
    # Total Tasks Card
    ws_dash.merge_cells('B5:C5')
    ws_dash['B5'] = "TOTAL TUGAS KELOMPOK"
    ws_dash['B5'].font = Font(name=font_family, size=9, bold=True, color="6B4F3A")
    ws_dash['B5'].alignment = align_center
    ws_dash['B5'].fill = card_fill
    
    ws_dash.merge_cells('B6:C7')
    ws_dash['B6'] = "=COUNTA('Gantt Chart Schedule'!B12:B50)"
    ws_dash['B6'].font = Font(name=font_family, size=24, bold=True, color="6B4F3A")
    ws_dash['B6'].alignment = align_center
    ws_dash['B6'].fill = card_fill
    
    # Completed Tasks Card
    ws_dash.merge_cells('E5:F5')
    ws_dash['E5'] = "TUGAS SELESAI"
    ws_dash['E5'].font = Font(name=font_family, size=9, bold=True, color="0F5132")
    ws_dash['E5'].alignment = align_center
    ws_dash['E5'].fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    
    ws_dash.merge_cells('E6:F7')
    ws_dash['E6'] = "=COUNTIF('Gantt Chart Schedule'!H12:H50, \"Completed\")"
    ws_dash['E6'].font = Font(name=font_family, size=24, bold=True, color="0F5132")
    ws_dash['E6'].alignment = align_center
    ws_dash['E6'].fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    
    # In Progress Card
    ws_dash.merge_cells('H5:I5')
    ws_dash['H5'] = "SEDANG DIKERJAKAN"
    ws_dash['H5'].font = Font(name=font_family, size=9, bold=True, color="664D03")
    ws_dash['H5'].alignment = align_center
    ws_dash['H5'].fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    ws_dash.merge_cells('H6:I7')
    ws_dash['H6'] = "=COUNTIF('Gantt Chart Schedule'!H12:H50, \"In Progress\")"
    ws_dash['H6'].font = Font(name=font_family, size=24, bold=True, color="664D03")
    ws_dash['H6'].alignment = align_center
    ws_dash['H6'].fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Border for cards
    for r in range(5, 8):
        for c in [2, 3]:
            ws_dash.cell(row=r, column=c).border = thin_border
        for c in [5, 6]:
            ws_dash.cell(row=r, column=c).border = thin_border
        for c in [8, 9]:
            ws_dash.cell(row=r, column=c).border = thin_border
            
    # Project progress bar
    ws_dash['B9'] = "TOTAL PROGRESS PROYEK:"
    ws_dash['B9'].font = section_font
    ws_dash['E9'] = "=AVERAGE('Gantt Chart Schedule'!G12:G50)"
    ws_dash['E9'].font = Font(name=font_family, size=12, bold=True, color="6B4F3A")
    ws_dash['E9'].number_format = '0.0%'
    ws_dash['E9'].alignment = align_left
    ws_dash['E9'].border = Border(bottom=Side(style="medium", color="6B4F3A"))

    # Table 2: Beban Kerja per Anggota Kelompok
    ws_dash['B12'] = "BEBAN KERJA & PROGRESS ANGGOTA KELOMPOK"
    ws_dash['B12'].font = section_font
    
    # Headers
    ws_dash['B14'] = "Nama Anggota (PIC)"
    ws_dash['B14'].font = header_font
    ws_dash['B14'].fill = primary_header_fill
    ws_dash['B14'].alignment = align_center
    ws_dash['B14'].border = thin_border
    
    ws_dash['C14'] = "Jumlah Tugas"
    ws_dash['C14'].font = header_font
    ws_dash['C14'].fill = primary_header_fill
    ws_dash['C14'].alignment = align_center
    ws_dash['C14'].border = thin_border
    
    ws_dash['E14'] = "Rata-rata Progress"
    ws_dash['E14'].font = header_font
    ws_dash['E14'].fill = primary_header_fill
    ws_dash['E14'].alignment = align_center
    ws_dash['E14'].border = thin_border
    
    pics = ["Bariandono", "Ajie", "Sarah", "Dewi"]
    row_start = 15
    for idx, pic_name in enumerate(pics):
        r = row_start + idx
        
        c1 = ws_dash.cell(row=r, column=2, value=pic_name)
        c1.font = data_bold_font
        c1.alignment = align_left
        c1.border = thin_border
        
        c2 = ws_dash.cell(row=r, column=3, value=f"=COUNTIF('Gantt Chart Schedule'!C$12:C$50, B{r})")
        c2.font = data_font
        c2.alignment = align_center
        c2.border = thin_border
        
        c3 = ws_dash.cell(row=r, column=5, value=f"=IF(C{r}=0, 0, AVERAGEIFS('Gantt Chart Schedule'!G$12:G$50, 'Gantt Chart Schedule'!C$12:C$50, B{r}))")
        c3.font = data_font
        c3.number_format = '0.0%'
        c3.alignment = align_center
        c3.border = thin_border
        
        # Zebra striping
        if idx % 2 == 1:
            c1.fill = zebra_fill
            c2.fill = zebra_fill
            c3.fill = zebra_fill
            
    # ---------------------------------------------------------
    # SHEET 3: GANTT CHART SCHEDULE
    # ---------------------------------------------------------
    ws_sched = wb.create_sheet(title="Gantt Chart Schedule")
    ws_sched.views.sheetView[0].showGridLines = True
    
    # Base columns width setting
    ws_sched.column_dimensions['A'].width = 5   # ID
    ws_sched.column_dimensions['B'].width = 30  # Task Name
    ws_sched.column_dimensions['C'].width = 15  # PIC
    ws_sched.column_dimensions['D'].width = 12  # Start Date
    ws_sched.column_dimensions['E'].width = 12  # End Date
    ws_sched.column_dimensions['F'].width = 9   # Duration
    ws_sched.column_dimensions['G'].width = 11  # Progress
    ws_sched.column_dimensions['H'].width = 13  # Status
    ws_sched.column_dimensions['I'].width = 3   # Spacer
    
    # Project Settings (Top Left Block)
    ws_sched['B2'] = "JADWAL TUGAS & GANTT CHART KELOMPOK"
    ws_sched['B2'].font = title_font
    
    ws_sched['B4'] = "Project Start Date:"
    ws_sched['B4'].font = Font(name=font_family, size=10, bold=True, color="6B4F3A")
    ws_sched['B4'].alignment = align_left
    
    # Set default date in C4 (Excel serial date or date string)
    ws_sched['C4'] = "2026-06-15"
    ws_sched['C4'].font = Font(name=font_family, size=10, bold=True, color="333333")
    ws_sched['C4'].alignment = align_center
    ws_sched['C4'].border = thin_border
    ws_sched['C4'].fill = card_fill
    
    ws_sched['B5'] = "Total Project Duration:"
    ws_sched['B5'].font = note_font
    
    ws_sched['C5'] = "=MAX(E12:E50)-MIN(D12:D50)+1"
    ws_sched['C5'].font = note_font
    ws_sched['C5'].alignment = align_center
    
    # Calendar timeline headers (Row 9 - 10, Columns J to AM - 30 days)
    start_col = 10  # 'J'
    num_days = 30
    
    # Setup day columns dimensions
    for d_idx in range(num_days):
        col_let = get_column_letter(start_col + d_idx)
        ws_sched.column_dimensions[col_let].width = 4
        
        # Day index (Row 9)
        c_day = ws_sched.cell(row=9, column=start_col + d_idx)
        c_day.value = f"H-{d_idx+1}"
        c_day.font = Font(name=font_family, size=8, bold=True, color="8A6D55")
        c_day.alignment = align_center
        c_day.border = thin_border
        
        # Dynamic dates (Row 10)
        c_date = ws_sched.cell(row=10, column=start_col + d_idx)
        if d_idx == 0:
            c_date.value = "=$C$4"
        else:
            prev_col_let = get_column_letter(start_col + d_idx - 1)
            c_date.value = f"={prev_col_let}10+1"
            
        c_date.font = Font(name=font_family, size=8, color="FFFFFF")
        c_date.fill = sub_header_fill
        c_date.number_format = 'dd/mm'
        c_date.alignment = align_vertical_header
        c_date.border = thin_border
        
    # Table headers (Row 11)
    headers = [
        ("A11", "ID"),
        ("B11", "Nama Tugas / Kegiatan"),
        ("C11", "PIC (Anggota)"),
        ("D11", "Start Date"),
        ("E11", "End Date"),
        ("F11", "Durasi"),
        ("G11", "Progress"),
        ("H11", "Status")
    ]
    
    for cell_id, text in headers:
        cell = ws_sched[cell_id]
        cell.value = text
        cell.font = header_font
        cell.fill = primary_header_fill
        cell.alignment = align_center
        cell.border = header_border
        
    # Bottom line border for timeline spacer column row 11
    ws_sched.cell(row=11, column=9).border = Border(bottom=thick_border_side)
    for d_idx in range(num_days):
        ws_sched.cell(row=11, column=start_col + d_idx).border = Border(bottom=thick_border_side)
        
    # Sample Tasks Data
    sample_tasks = [
        ("1", "Riset Topik & Cari Masalah", "Bariandono", "2026-06-15", "2026-06-18", 1.0),
        ("2", "Desain Outline Proyek", "Ajie", "2026-06-19", "2026-06-21", 0.8),
        ("3", "Pengumpulan Data Kuesioner", "Sarah", "2026-06-22", "2026-06-27", 0.3),
        ("4", "Olahan Data SPSS/Stata", "Dewi", "2026-06-28", "2026-06-30", 0.0),
        ("5", "Penyusunan Draf Bab I-III", "Bariandono", "2026-07-01", "2026-07-05", 0.0),
        ("6", "Redesain PPT Sidang Sempro", "Ajie", "2026-07-06", "2026-07-10", 0.0),
        ("7", "Review Plagiasi & Turnitin", "Sarah", "2026-07-11", "2026-07-13", 0.0),
        ("8", "Finalisasi & Pendaftaran Sidang", "Dewi", "2026-07-14", "2026-07-15", 0.0)
    ]
    
    # Populate Tasks & Formulas
    max_tasks_rows = 40
    for i in range(max_tasks_rows):
        r = 12 + i
        has_sample = i < len(sample_tasks)
        
        # ID
        c_id = ws_sched.cell(row=r, column=1, value=sample_tasks[i][0] if has_sample else (i+1 if i < 15 else ""))
        c_id.font = data_bold_font
        c_id.alignment = align_center
        c_id.border = thin_border
        
        # Task Name
        c_name = ws_sched.cell(row=r, column=2, value=sample_tasks[i][1] if has_sample else "")
        c_name.font = data_font
        c_name.alignment = align_left
        c_name.border = thin_border
        
        # PIC
        c_pic = ws_sched.cell(row=r, column=3, value=sample_tasks[i][2] if has_sample else "")
        c_pic.font = data_font
        c_pic.alignment = align_center
        c_pic.border = thin_border
        
        # Start Date
        c_start = ws_sched.cell(row=r, column=4, value=sample_tasks[i][3] if has_sample else "")
        c_start.font = data_font
        c_start.number_format = 'yyyy-mm-dd'
        c_start.alignment = align_center
        c_start.border = thin_border
        
        # End Date
        c_end = ws_sched.cell(row=r, column=5, value=sample_tasks[i][4] if has_sample else "")
        c_end.font = data_font
        c_end.number_format = 'yyyy-mm-dd'
        c_end.alignment = align_center
        c_end.border = thin_border
        
        # Duration (Formula)
        c_dur = ws_sched.cell(row=r, column=6, value=f"=IF(OR(ISBLANK(D{r}), ISBLANK(E{r})), \"\", E{r}-D{r}+1)")
        c_dur.font = data_font
        c_dur.alignment = align_center
        c_dur.border = thin_border
        
        # Progress %
        c_prog = ws_sched.cell(row=r, column=7, value=sample_tasks[i][5] if has_sample else "")
        c_prog.font = data_font
        c_prog.number_format = '0%'
        c_prog.alignment = align_center
        c_prog.border = thin_border
        
        # Status (Formula)
        c_status = ws_sched.cell(row=r, column=8, value=f"=IF(ISBLANK(B{r}), \"\", IF(G{r}=0, \"Not Started\", IF(G{r}=1, \"Completed\", \"In Progress\")))")
        c_status.font = data_bold_font
        c_status.alignment = align_center
        c_status.border = thin_border
        
        # Spacer column (Column I)
        ws_sched.cell(row=r, column=9).border = Border(left=thin_border_side)
        
        # Gantt Chart grid cell formulas (Column J to AM)
        for d_idx in range(num_days):
            c_gantt_col = start_col + d_idx
            col_let = get_column_letter(c_gantt_col)
            c_gantt = ws_sched.cell(row=r, column=c_gantt_col)
            # Renders '█' if the date falls in the range, else empty string.
            # This makes the Gantt bar print well and remain visible in basic viewers.
            c_gantt.value = f"=IF(AND(NOT(ISBLANK(D{r})), {col_let}$10>=D{r}, {col_let}$10<=E{r}), \"█\", \"\")"
            c_gantt.font = Font(name=font_family, size=10, color="E8C9A0") # Font color same as Gantt fill (Sand Gold)
            c_gantt.alignment = align_center
            c_gantt.border = thin_border
            
        # Zebra striping on task table
        if i % 2 == 1:
            for col_idx in range(1, 9):
                ws_sched.cell(row=r, column=col_idx).fill = zebra_fill
                
    # ---------------------------------------------------------
    # CONDITIONAL FORMATTING RULES
    # ---------------------------------------------------------
    # 1. Gantt grid bar color (Sand Gold highlight for '█' cells)
    gantt_range = f"J12:{get_column_letter(start_col + num_days - 1)}{12 + max_tasks_rows - 1}"
    gantt_fill = PatternFill(start_color="E8C9A0", end_color="E8C9A0", fill_type="solid")
    gantt_font = Font(name=font_family, size=10, color="E8C9A0", bold=True)
    ws_sched.conditional_formatting.add(gantt_range, CellIsRule(operator='equal', formula=['"█"'], fill=gantt_fill, font=gantt_font))
    
    # 2. Status column color codes
    status_range = f"H12:H{12 + max_tasks_rows - 1}"
    
    # Completed (Soft Green)
    comp_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    comp_font = Font(name=font_family, size=10, bold=True, color="0F5132")
    ws_sched.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Completed"'], fill=comp_fill, font=comp_font))
    
    # In Progress (Soft Yellow)
    prog_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    prog_font = Font(name=font_family, size=10, bold=True, color="664D03")
    ws_sched.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"In Progress"'], fill=prog_fill, font=prog_font))
    
    # Not Started (Light Grey)
    ns_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    ns_font = Font(name=font_family, size=10, bold=True, color="6C757D")
    ws_sched.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Not Started"'], fill=ns_fill, font=ns_font))

    # Save workbook
    output_xlsx = 'projects/project-management-excel/Gantt_Chart_Kelompok.xlsx'
    wb.save(output_xlsx)
    print(f"Workbook successfully generated: {output_xlsx}")

if __name__ == "__main__":
    create_gantt_chart()
